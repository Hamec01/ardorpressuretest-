from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from wika_report import __version__
from wika_report.models import AnalysisResult, ColumnMapping, FileMetadata, CleaningStats


def build_excel_report(
    df_raw: pd.DataFrame,
    df_clean: pd.DataFrame,
    meta: FileMetadata,
    mapping: ColumnMapping,
    stats: CleaningStats,
    analysis: AnalysisResult,
    graph_path: Optional[Path],
    output_path: Path
) -> Path:
    """
    Creates a structured Excel report (.xlsx) with 4 sheets: Summary, Cleaned Data, Raw Data, Metadata.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=14, bold=True, color="1F4E79")
    bold_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    
    pass_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    pass_font = Font(name=font_family, size=11, bold=True, color="274E13")
    
    fail_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    fail_font = Font(name=font_family, size=11, bold=True, color="783F04")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # ==========================================
    # SHEET 1: Summary
    # ==========================================
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum["A1"] = "WIKA CPG1500 PRESSURE ANALYSIS REPORT"
    ws_sum["A1"].font = title_font

    summary_data = [
        ("Data File:", meta.input_path.name),
        ("Device:", meta.device_info.get("Device", meta.device_info.get("Model", "WIKA CPG1500"))),
        ("Serial No / Wika Nr:", analysis.custom_meta.wika_nr or meta.device_info.get("Serial No", meta.device_info.get("SN", "N/A"))),
        ("Detected Input Unit:", meta.detected_unit or "Undetermined"),
        ("Target Unit:", "bar"),
        ("Start Date/Time:", analysis.start_time.strftime("%Y-%m-%d %H:%M:%S") if analysis.start_time else "N/A"),
        ("End Date/Time:", analysis.end_time.strftime("%Y-%m-%d %H:%M:%S") if analysis.end_time else "N/A"),
        ("Duration:", analysis.duration_formatted),
        ("Test Pressure:", analysis.custom_meta.test_pressure or "N/A"),
        ("System:", analysis.custom_meta.system or "N/A"),
        ("Log No:", analysis.custom_meta.log_no or "N/A"),
        ("Ins No:", analysis.custom_meta.ins_no or "N/A"),
        ("Project:", analysis.custom_meta.project or "N/A"),
        ("Note:", analysis.custom_meta.note or "N/A"),
        ("Pipe Logs:", analysis.custom_meta.pipe_logs_text.replace("\n", ", ") if getattr(analysis.custom_meta, "pipe_logs_text", None) else "N/A"),
        ("Raw CSV Rows:", analysis.total_raw_rows),
        ("Valid Points:", analysis.valid_points_count),
        ("Excluded Rows:", analysis.excluded_rows_count),
        ("Start Pressure (bar):", round(analysis.start_pressure_bar, 5)),
        ("End Pressure (bar):", round(analysis.end_pressure_bar, 5)),
        ("Min Pressure (bar):", round(analysis.min_pressure_bar, 5)),
        ("Max Pressure (bar):", round(analysis.max_pressure_bar, 5)),
        ("Mean Pressure (bar):", round(analysis.mean_pressure_bar, 5)),
        ("Median Pressure (bar):", round(analysis.median_pressure_bar, 5)),
        ("Std Deviation (bar):", round(analysis.std_pressure_bar, 5)),
        ("Total Delta (bar):", round(analysis.total_delta_bar, 5)),
        ("Pressure Range (bar):", round(analysis.range_bar, 5)),
        ("Max Pressure Time:", analysis.max_pressure_time_str),
        ("Max Rise Rate (bar/min):", round(analysis.max_rise_rate_bar_per_min, 4)),
        ("Max Drop Rate (bar/min):", round(analysis.max_drop_rate_bar_per_min, 4)),
    ]


    ws_sum.cell(row=3, column=1, value="Parameter").font = header_font
    ws_sum.cell(row=3, column=1).fill = header_fill
    ws_sum.cell(row=3, column=2, value="Value").font = header_font
    ws_sum.cell(row=3, column=2).fill = header_fill

    row_idx = 4
    for label, val in summary_data:
        c1 = ws_sum.cell(row=row_idx, column=1, value=label)
        c2 = ws_sum.cell(row=row_idx, column=2, value=val)
        c1.font = bold_font
        c2.font = normal_font
        c1.border = thin_border
        c2.border = thin_border
        row_idx += 1

    # PASS / FAIL evaluation block
    if analysis.hold_stats and analysis.hold_stats.enabled:
        row_idx += 1
        ws_sum.cell(row=row_idx, column=1, value="TEST RESULT:").font = bold_font
        c_res = ws_sum.cell(row=row_idx, column=2, value=analysis.hold_stats.status)
        if analysis.hold_stats.status == "PASS":
            c_res.fill = pass_fill
            c_res.font = pass_font
        elif analysis.hold_stats.status == "FAIL":
            c_res.fill = fail_fill
            c_res.font = fail_font

        if analysis.hold_stats.fail_reasons:
            row_idx += 1
            ws_sum.cell(row=row_idx, column=1, value="Non-compliance Reasons:").font = bold_font
            reasons_str = "; ".join(analysis.hold_stats.fail_reasons)
            ws_sum.cell(row=row_idx, column=2, value=reasons_str).font = normal_font

    # Embed graph image into Summary sheet
    if graph_path and graph_path.exists():
        img = OpenPyxlImage(str(graph_path))
        img.width = 750
        img.height = 375
        ws_sum.add_image(img, "D3")

    # ==========================================
    # SHEET 2: Cleaned Data
    # ==========================================
    ws_clean = wb.create_sheet(title="Cleaned Data")
    ws_clean.views.sheetView[0].showGridLines = True

    clean_headers = [
        "Point No", "Original Time", "Time from Start (sec)",
        "Time from Start (min)", "Original Pressure", "Original Unit", "Pressure (bar)"
    ]

    for col_i, h_text in enumerate(clean_headers, 1):
        cell = ws_clean.cell(row=1, column=col_i, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r_i, row in enumerate(df_clean.itertuples(), 2):
        ws_clean.cell(row=r_i, column=1, value=getattr(row, "point_idx", r_i - 1)).alignment = Alignment(horizontal="center")
        ws_clean.cell(row=r_i, column=2, value=str(getattr(row, "_time_raw_str", "")))
        ws_clean.cell(row=r_i, column=3, value=round(float(getattr(row, "time_rel_sec", 0.0)), 2)).number_format = "0.00"
        ws_clean.cell(row=r_i, column=4, value=round(float(getattr(row, "time_rel_min", 0.0)), 3)).number_format = "0.000"
        ws_clean.cell(row=r_i, column=5, value=str(getattr(row, "_raw_pressure", "")))
        ws_clean.cell(row=r_i, column=6, value=meta.detected_unit or "")
        ws_clean.cell(row=r_i, column=7, value=float(getattr(row, "pressure_bar", 0.0))).number_format = "0.00000"

    ws_clean.freeze_panes = "A2"
    ws_clean.auto_filter.ref = ws_clean.dimensions

    # ==========================================
    # SHEET 3: Raw Data
    # ==========================================
    ws_raw = wb.create_sheet(title="Raw Data")
    ws_raw.views.sheetView[0].showGridLines = True

    for col_i, h_text in enumerate(df_raw.columns, 1):
        cell = ws_raw.cell(row=1, column=col_i, value=str(h_text))
        cell.font = header_font
        cell.fill = header_fill

    for r_i, row_vals in enumerate(df_raw.values, 2):
        for c_i, val in enumerate(row_vals, 1):
            ws_raw.cell(row=r_i, column=c_i, value=str(val) if pd.notna(val) else "")

    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = ws_raw.dimensions

    # ==========================================
    # SHEET 4: Metadata
    # ==========================================
    ws_meta = wb.create_sheet(title="Metadata")
    ws_meta.views.sheetView[0].showGridLines = True

    meta_rows = [
        ("Parser Parameter", "Value"),
        ("File Encoding:", meta.encoding),
        ("Column Delimiter:", f"'{meta.delimiter}'"),
        ("Decimal Separator:", f"'{meta.decimal_sep}'"),
        ("Header Line (Index):", meta.header_line_idx + 1),
        ("Detected Time Column:", mapping.time_col or "Not found"),
        ("Detected Date Column:", mapping.date_col or "Not found"),
        ("Detected Pressure Column:", mapping.pressure_col or "Not found"),
        ("Detected Input Unit:", meta.detected_unit or "Undetermined"),
        ("Unit Detection Source:", meta.detected_unit_source or "N/A"),
        ("Total Raw CSV Rows:", stats.total_raw_rows),
        ("Accepted Points:", stats.clean_rows),
        ("Excluded Rows:", stats.excluded_rows),
        ("Program Version:", f"wika-cpg1500-graph v{__version__}"),
        ("Report Generation Time:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for r_i, (k, v) in enumerate(meta_rows, 1):
        c1 = ws_meta.cell(row=r_i, column=1, value=k)
        c2 = ws_meta.cell(row=r_i, column=2, value=v)
        if r_i == 1:
            c1.font = header_font
            c1.fill = header_fill
            c2.font = header_font
            c2.fill = header_fill
        else:
            c1.font = bold_font
            c2.font = normal_font
            c1.border = thin_border
            c2.border = thin_border

    if stats.exclusion_reasons:
        r_start = len(meta_rows) + 2
        ws_meta.cell(row=r_start, column=1, value="Exclusion Reason").font = header_font
        ws_meta.cell(row=r_start, column=1).fill = header_fill
        ws_meta.cell(row=r_start, column=2, value="Count").font = header_font
        ws_meta.cell(row=r_start, column=2).fill = header_fill

        for r_i, (reason, count) in enumerate(stats.exclusion_reasons.items(), r_start + 1):
            c1 = ws_meta.cell(row=r_i, column=1, value=reason)
            c2 = ws_meta.cell(row=r_i, column=2, value=count)
            c1.font = normal_font
            c2.font = normal_font
            c1.border = thin_border
            c2.border = thin_border

    # Auto-fit column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 30

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
